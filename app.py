import os
import sys
import json
import sqlite3
import datetime
import threading
import time
import re
import hashlib
import logging
from flask import Flask, request, jsonify, send_from_directory, render_template_string, Response
from flask_cors import CORS
from werkzeug.utils import secure_filename
import requests
from bs4 import BeautifulSoup
import pandas as pd

# 日志配置
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ============================================================
# 药房进货比较系统 - Flask后端主入口
# 部署平台: VicroCode
# Python版本: 3.10
# ============================================================

# 路径处理：兼容 PyInstaller 打包环境
def _get_resource_path(relative_path):
    """获取资源文件路径（兼容开发环境和打包后的环境）"""
    if getattr(sys, 'frozen', False):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), relative_path)

# 创建Flask应用 - 混合项目配置
app = Flask(__name__,
            static_folder=_get_resource_path('static'),
            static_url_path='',
            template_folder=_get_resource_path('templates'))

# 密钥配置
app.secret_key = 'pharmacy-system-secret-key-2024'

# CORS配置 - 必须包含所有VicroCode域名
CORS(app, resources={
    r"/*": {
        "origins": [
            "http://localhost:5173",
            "https://www.vicrocode.com",
            "https://vicrocode.com",
            "https://www.vioco.cn",
            "https://vioco.cn",
            "https://www.vicoco.cn",
            "https://vicoco.cn",
            "https://www.vicrocode.cn",
            "https://vicrocode.cn"
        ],
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization", "X-User-Id", "X-Username"],
        "supports_credentials": True
    }
})

# 移除X-Frame-Options，允许iframe显示
@app.after_request
def add_security_headers(response):
    response.headers.pop('X-Frame-Options', None)
    origin = request.headers.get('Origin')
    if origin:
        response.headers['Access-Control-Allow-Origin'] = origin
    response.headers['Access-Control-Allow-Credentials'] = 'true'
    return response

# ============================================================
# 数据库配置
# ============================================================
DATABASE = 'pharmacy.db'

def get_db():
    """获取数据库连接"""
    db = sqlite3.connect(DATABASE)
    db.row_factory = sqlite3.Row
    return db

def init_db():
    """初始化数据库"""
    db = get_db()
    cursor = db.cursor()
    
    # 供应商表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS suppliers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            url TEXT NOT NULL,
            account TEXT NOT NULL,
            password TEXT NOT NULL,
            status INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # 药品搜索记录表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS search_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            barcode TEXT,
            product_name TEXT,
            specification TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # 价格比较结果表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS price_results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            search_record_id INTEGER,
            supplier_id INTEGER,
            supplier_name TEXT,
            product_name TEXT,
            specification TEXT,
            price REAL,
            validity_date TEXT,
            stock_status TEXT,
            direct_link TEXT,
            search_status TEXT DEFAULT 'pending',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (search_record_id) REFERENCES search_records(id)
        )
    ''')
    
    # 初始化默认供应商数据
    default_suppliers = [
        ('俊龙', 'http://shop.szsjlyy.com/', '48884', 'YXGS888'),
        ('康之源', 'https://www.gdkzyyy.com/', '13725533203', '533203'),
        ('中源', 'http://shop.szszyyy.com/', 'DD423', '123456'),
        ('明华堂', 'https://web.szmhtyy.com/', '13725533203', '123456')
    ]
    
    cursor.execute("SELECT COUNT(*) FROM suppliers")
    if cursor.fetchone()[0] == 0:
        for supplier in default_suppliers:
            cursor.execute('''
                INSERT INTO suppliers (name, url, account, password)
                VALUES (?, ?, ?, ?)
            ''', supplier)
    
    db.commit()
    db.close()

# 初始化数据库
init_db()

# ============================================================
# 爬虫模块 - 通过API获取供应商数据
# 俊龙: 86yqy.com平台 JWT Bearer Token认证
# 庆丰裕系统(康之源/中源/明华堂): web_前缀API + token认证
# ============================================================

class PharmacyCrawler:
    """药房供应商爬虫类 - 使用真实API获取数据"""

    # 俊龙API配置
    JUNLONG_API = 'https://swoole.86yqy.com'
    JUNLONG_CLIENT_ID = 2
    JUNLONG_CLIENT_SECRET = '2sGHkOgJa4fam6dYb7s5Wl3DeppmVUA4KCZIxNAI'
    JUNLONG_CREDENTIALS = {'username': '48884', 'password': 'YXGS888'}
    JUNLONG_COMPANY_ID = 6255

    # 庆丰裕系统配置 (康之源/中源/明华堂)
    QFY_SYSTEMS = {
        '康之源': {
            'web_url': 'https://www.gdkzyyy.com',
            'api_url': 'https://api.gdkzyyy.com',
            'username': '13725533203',
            'password': '533203',
        },
        '中源': {
            'web_url': 'http://shop.szszyyy.com',
            'api_url': 'http://api.szszyyy.com',
            'username': 'DD423',
            'password': '123456',
        },
        '明华堂': {
            'web_url': 'https://web.szmhtyy.com',
            'api_url': 'https://api.szmhtyy.com',
            'username': '13725533203',
            'password': '123456',
        },
    }

    def __init__(self):
        self.session = requests.Session()
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
        }
        self.session.headers.update(self.headers)
        # Token缓存 (类级别，避免重复登录)
        self._junlong_token = None
        self._junlong_token_time = 0
        self._qfy_tokens = {}  # {name: {'token': xx, 'time': xx}}

    # ===================== 俊龙系统 =====================

    def _get_junlong_token(self):
        """获取俊龙JWT Token，带缓存"""
        now = time.time()
        # Token有效期约120天(10367882秒)，缓存1小时
        if self._junlong_token and (now - self._junlong_token_time) < 3600:
            return self._junlong_token

        try:
            # 使用独立session避免header冲突
            junlong_session = requests.Session()
            junlong_session.headers.update({
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Accept': 'application/json, text/plain, */*',
                'Content-Type': 'application/json;charset=UTF-8',
                'Origin': 'http://shop.szsjlyy.com',
                'Referer': 'http://shop.szsjlyy.com/',
            })
            login_url = f'{self.JUNLONG_API}/api/authorizations'
            login_data = {
                'username': self.JUNLONG_CREDENTIALS['username'],
                'password': self.JUNLONG_CREDENTIALS['password'],
                'client_id': self.JUNLONG_CLIENT_ID,
                'client_secret': self.JUNLONG_CLIENT_SECRET,
                'grant_type': 'password',
                'provider': 'api'
            }
            resp = junlong_session.post(login_url, json=login_data, timeout=15)
            if resp.status_code in (200, 201):
                result = resp.json()
                self._junlong_token = result.get('access_token', '')
                self._junlong_token_time = now
                logger.info("俊龙登录成功，获取token")
                return self._junlong_token
            else:
                logger.error(f"俊龙登录失败: {resp.status_code} {resp.text[:200]}")
                return None
        except Exception as e:
            logger.error(f"俊龙登录异常: {e}")
            return None

    def search_junlong(self, barcode):
        """俊龙药房 - 通过86yqy.com API搜索"""
        try:
            token = self._get_junlong_token()
            if not token:
                return self._error_result('俊龙', barcode, '登录失败')

            # 使用独立session
            junlong_session = requests.Session()
            junlong_session.headers.update({
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Accept': 'application/json, text/plain, */*',
                'Authorization': f'Bearer {token}',
                'Referer': 'http://shop.szsjlyy.com/',
            })
            api_url = f'{self.JUNLONG_API}/api/shop/product/list'
            params = {'keyword': barcode, 'page': 1, 'pageSize': 16}

            resp = junlong_session.get(api_url, params=params, timeout=15)
            if resp.status_code != 200:
                return self._error_result('俊龙', barcode, f'API错误 {resp.status_code}')

            result = resp.json()
            items = result.get('data', [])
            total = result.get('total', 0)

            if total == 0 or not items:
                return {
                    'status': 'not_found',
                    'product_name': '没有',
                    'specification': '-',
                    'price': None,
                    'original_price': None,
                    'discount_price': None,
                    'validity_date': '-',
                    'stock_status': '-',
                    'direct_link': 'http://shop.szsjlyy.com/category?keyword=' + barcode,
                    'product_image': '',
                    'manufacturer': '',
                    'cart_quantity': None,
                }

            # 取第一个匹配商品
            item = items[0]
            stock_detail = item.get('stock_detail') or {}
            # 获取价格 - 俊龙所有商品都无折扣价，无需爬取优惠价
            price = item.get('whole_price')  # 批发价/实际售价
            try:
                price = float(price) if price else None
            except (ValueError, TypeError):
                price = None

            # 俊龙无优惠：原价就是售价，折后价永远为 None
            original_price = price
            discount_price = None

            # 获取商品图片 - 修复: 俊龙图片可能以 // 开头，需拼接 https:
            product_image = item.get('image') or item.get('thumb') or ''
            if not product_image:
                _images = item.get('images')
                if isinstance(_images, list) and _images:
                    product_image = _images[0]
                elif _images:
                    product_image = str(_images)
            if product_image and not product_image.startswith('http'):
                product_image = 'https:' + product_image

            # 获取生产厂家 - 实际API使用 factory_name
            manufacturer = item.get('factory_name') or item.get('producer') or ''

            # 缓存 sku_id 用于后续购物车查询
            sku_list = item.get('sku', [])
            if sku_list and len(sku_list) > 0:
                sku_id = sku_list[0].get('id', item.get('id'))
            else:
                sku_id = item.get('id')

            return {
                'status': 'found',
                'product_name': item.get('product_name') or item.get('standard_name') or '没有',
                'specification': item.get('attr') or '-',
                'price': price,
                'original_price': original_price,
                'discount_price': discount_price,
                'validity_date': stock_detail.get('expiry_date') or '-',
                'stock_status': item.get('stock_show') or ('有货' if (item.get('stock') or 0) > 0 else '缺货'),
                'direct_link': 'http://shop.szsjlyy.com/category?keyword=' + barcode,
                'product_image': product_image or '',
                'manufacturer': manufacturer,
                'cart_quantity': None,
                '_sku_id': sku_id,  # 内部字段，用于购物车查询
            }

        except Exception as e:
            logger.error(f"俊龙搜索异常: {e}")
            return self._error_result('俊龙', barcode, str(e))

    # ===================== 庆丰裕系统 =====================

    def _get_qfy_token(self, name):
        """获取庆丰裕系统token，带缓存"""
        now = time.time()
        cached = self._qfy_tokens.get(name)
        # Token有效期约8天，缓存1小时
        if cached and (now - cached.get('time', 0)) < 3600:
            return cached['token']

        config = self.QFY_SYSTEMS.get(name)
        if not config:
            return None

        try:
            session = requests.Session()
            session.headers.update(self.headers)
            session.headers['Referer'] = f"{config['web_url']}/member/sign/login.html"

            # 访问首页获取cookie
            session.get(config['web_url'] + '/', timeout=10)

            # 登录 - 密码MD5加密
            password_md5 = hashlib.md5(config['password'].encode()).hexdigest()
            login_url = f"{config['api_url']}/web_login.html"
            login_data = {
                'user_phone': config['username'],
                'login_pwd': password_md5,
                'time': str(int(now * 1000))
            }
            resp = session.post(login_url, data=login_data, timeout=10)
            result = resp.json()

            if result.get('status') == 1:
                token = result['token']['token']
                self._qfy_tokens[name] = {'token': token, 'time': now, 'session': session}
                logger.info(f"{name}登录成功")
                return token
            else:
                logger.error(f"{name}登录失败: {result.get('msg')}")
                return None
        except Exception as e:
            logger.error(f"{name}登录异常: {e}")
            return None

    def _search_qfy(self, name, barcode):
        """庆丰裕系统通用搜索方法"""
        config = self.QFY_SYSTEMS.get(name)
        if not config:
            return self._error_result(name, barcode, '配置缺失')

        try:
            token = self._get_qfy_token(name)
            if not token:
                return self._error_result(name, barcode, '登录失败')

            # 使用缓存的session或新建
            cached = self._qfy_tokens.get(name, {})
            sess = cached.get('session') or self.session

            api_url = f"{config['api_url']}/web_goods_list.html"
            goods_data = {
                'token': token,
                'time': str(int(time.time() * 1000)),
                'param[word]': barcode,
                'order[val]': '1',
                'order[type]': 'desc',
                'hasqty': '0',
                'hasmedicare': '0',
                'buyhistory': '0',
                'hasstore': '0',
                'page': '1'
            }
            resp = sess.post(api_url, data=goods_data, timeout=15)
            result = resp.json()

            if result.get('status') != 1:
                # Token可能过期，清除缓存重试一次
                self._qfy_tokens.pop(name, None)
                token = self._get_qfy_token(name)
                if not token:
                    return self._error_result(name, barcode, '登录失败')
                goods_data['token'] = token
                goods_data['time'] = str(int(time.time() * 1000))
                resp = sess.post(api_url, data=goods_data, timeout=15)
                result = resp.json()

            if result.get('status') != 1:
                return self._error_result(name, barcode, result.get('msg', 'API错误'))

            items = result.get('data', {}).get('data', [])
            if not items:
                return {
                    'status': 'not_found',
                    'product_name': '没有',
                    'specification': '-',
                    'price': None,
                    'original_price': None,
                    'discount_price': None,
                    'validity_date': '-',
                    'stock_status': '-',
                    'direct_link': f"{config['web_url']}/product/list.html?word={barcode}",
                    'product_image': '',
                    'manufacturer': '',
                    'cart_quantity': None,
                }

            item = items[0]
            # 获取价格 - 实际API只有price字段
            price = item.get('price')
            try:
                price = float(price) if price is not None else None
            except (ValueError, TypeError):
                price = None

            # 庆丰裕系统的优惠通过 rateprice 字段判断
            # rateprice > 0 且 < price 时有优惠，rateprice 就是折后价
            original_price = None
            discount_price = None

            rateprice = item.get('rateprice')
            try:
                rateprice = float(rateprice) if rateprice else 0
            except (ValueError, TypeError):
                rateprice = 0

            if rateprice > 0 and price and rateprice < price:
                # 有优惠：原价 = price，折后价 = rateprice
                original_price = price
                discount_price = rateprice
            else:
                # 无优惠：原价就是售价，折后价为 None
                original_price = price
                discount_price = None

            stock_qty = item.get('zqty', 0) or item.get('norqty', 0) or 0

            # 获取商品图片 - 庆丰裕系统图片处理
            # goods_thumb/minimg 返回的是不带扩展名的路径，需按供应商加后缀
            img = item.get('goods_thumb') or item.get('thumb') or item.get('image') or item.get('minimg') or ''
            if not img:
                _images = item.get('images')
                if isinstance(_images, list) and _images:
                    img = _images[0]
                elif _images:
                    img = str(_images)
            # 拼接完整URL：庆丰裕图片路径不带扩展名，需加后缀
            if img and not img.startswith('http'):
                if not img.endswith(('.jpg', '.jpeg', '.png', '.gif', '.webp')):
                    if name == '康之源':
                        img = config['api_url'] + img + '_thumb.jpg'
                    else:
                        img = config['api_url'] + img + '.jpg'
                else:
                    if img.startswith('/'):
                        img = config['api_url'] + img
                    else:
                        img = config['api_url'] + '/' + img

            # 获取生产厂家 - 修复: 优先 maker 再 producer 再 manufacturer
            manufacturer = item.get('maker') or item.get('producer') or item.get('manufacturer') or ''

            # 缓存 mid 用于后续购物车查询
            mid = item.get('id')

            return {
                'status': 'found',
                'product_name': item.get('goods_name') or '没有',
                'specification': item.get('specs') or '-',
                'price': price,
                'original_price': original_price,
                'discount_price': discount_price,
                'validity_date': item.get('out_date') or '-',
                'stock_status': '有货' if stock_qty > 0 else '缺货',
                'direct_link': f"{config['web_url']}/product/list.html?word={barcode}",
                'product_image': img or '',
                'manufacturer': manufacturer,
                'cart_quantity': None,
                '_mid': mid,  # 内部字段，用于购物车查询
            }

        except Exception as e:
            logger.error(f"{name}搜索异常: {e}")
            return self._error_result(name, barcode, str(e))

    def search_kangzhiyuan(self, barcode):
        """康之源 - 庆丰裕系统"""
        return self._search_qfy('康之源', barcode)

    def search_zhongyuan(self, barcode):
        """中源 - 庆丰裕系统"""
        return self._search_qfy('中源', barcode)

    def search_minghuatang(self, barcode):
        """明华堂 - 庆丰裕系统"""
        return self._search_qfy('明华堂', barcode)

    # ===================== 购物车数量内部查询方法 =====================

    def _get_cart_quantity_raw(self, supplier_name, barcode):
        """内部方法（不走HTTP），直接调用API获取购物车中某商品的数量
        用于搜索时同步填充 cart_quantity 字段
        """
        try:
            if supplier_name == '俊龙':
                return self._get_junlong_cart_quantity(barcode)
            elif supplier_name in self.QFY_SYSTEMS:
                return self._get_qfy_cart_quantity(supplier_name, barcode)
            else:
                return None
        except Exception as e:
            logger.warning(f"获取{supplier_name}购物车数量失败: {e}")
            return None

    def _get_junlong_cart_quantity(self, barcode):
        """俊龙购物车数量查询（内部方法）"""
        token = self._get_junlong_token()
        if not token:
            return None

        try:
            session = requests.Session()
            session.headers.update({
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                'Accept': 'application/json',
                'Authorization': f'Bearer {token}',
                'Referer': 'http://shop.szsjlyy.com/',
            })

            # 先搜索获取sku_id
            search_resp = session.get(
                f'{self.JUNLONG_API}/api/shop/product/list',
                params={'keyword': barcode, 'page': 1, 'pageSize': 5},
                timeout=15
            )
            search_result = search_resp.json()
            items = search_result.get('data', [])
            if not items:
                return 0

            item = items[0]
            sku_list = item.get('sku', [])
            sku_id = sku_list[0].get('id', item.get('id')) if sku_list else item.get('id')

            # 获取购物车列表
            cart_resp = session.get(
                f'{self.JUNLONG_API}/api/order/cart',
                params={'company_id': self.JUNLONG_COMPANY_ID},
                timeout=15
            )
            if cart_resp.status_code == 200:
                cart_data = cart_resp.json()
                # 数据结构: data[].data.valid_groups[].goods[]
                for group in cart_data.get('data', []):
                    inner_data = group.get('data', {})
                    valid_groups = inner_data.get('valid_groups', [])
                    for vg in valid_groups:
                        for g in vg.get('goods', []):
                            if str(g.get('sku_id', '')) == str(sku_id):
                                return g.get('quantity', 0)
            return 0
        except Exception as e:
            logger.warning(f"俊龙购物车数量查询失败: {e}")
            return None

    def _get_qfy_cart_quantity(self, supplier_name, barcode):
        """庆丰裕系统购物车数量查询（内部方法） - 使用 web_mycart.html"""
        config = self.QFY_SYSTEMS.get(supplier_name)
        if not config:
            return None

        token = self._get_qfy_token(supplier_name)
        if not token:
            return None

        try:
            cached = self._qfy_tokens.get(supplier_name, {})
            sess = cached.get('session') or self.session

            # 先搜索商品获取mid
            goods_data = {
                'token': token,
                'time': str(int(time.time() * 1000)),
                'param[word]': barcode,
                'order[val]': '1',
                'order[type]': 'desc',
                'hasqty': '0',
                'hasmedicare': '0',
                'buyhistory': '0',
                'hasstore': '0',
                'page': '1'
            }
            search_resp = sess.post(
                f"{config['api_url']}/web_goods_list.html",
                data=goods_data,
                timeout=15
            )
            search_result = search_resp.json()
            if search_result.get('status') != 1:
                return 0

            items = search_result.get('data', {}).get('data', [])
            if not items:
                return 0

            mid = items[0].get('id')

            # 获取购物车列表 - 修复: 使用 web_mycart.html（与 /api/cart/list 一致）
            cart_data_req = {
                'token': token,
                'time': str(int(time.time() * 1000)),
            }
            cart_resp = sess.post(
                f"{config['api_url']}/web_mycart.html",
                data=cart_data_req,
                timeout=15
            )
            try:
                cart_result = cart_resp.json()
                if cart_result.get('status') == 1:
                    # 遍历 normal[].list[] 查找匹配的商品
                    for group in cart_result.get('normal', []):
                        for item in group.get('list', []):
                            if str(item.get('mid', '')) == str(mid) or str(item.get('id', '')) == str(mid):
                                return item.get('buynum', 0)
                    # 也检查 invalid 列表
                    for item in cart_result.get('invalid', []):
                        if str(item.get('mid', '')) == str(mid) or str(item.get('id', '')) == str(mid):
                            return item.get('buynum', 0)
            except Exception as e:
                logger.warning(f"{supplier_name}购物车数量解析失败: {e}")
            return 0
        except Exception as e:
            logger.warning(f"{supplier_name}购物车数量查询失败: {e}")
            return None

    def _get_junlong_cart_qty_by_skuid(self, sku_id):
        """直接使用 sku_id 查询俊龙购物车数量"""
        token = self._get_junlong_token()
        if not token or not sku_id:
            return None
        try:
            session = requests.Session()
            session.headers.update({
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                'Accept': 'application/json',
                'Authorization': f'Bearer {token}',
                'Referer': 'http://shop.szsjlyy.com/',
            })
            cart_resp = session.get(
                f'{self.JUNLONG_API}/api/order/cart',
                params={'company_id': self.JUNLONG_COMPANY_ID},
                timeout=15
            )
            if cart_resp.status_code == 200:
                cart_data = cart_resp.json()
                for group in cart_data.get('data', []):
                    inner_data = group.get('data', {})
                    for vg in inner_data.get('valid_groups', []):
                        for g in vg.get('goods', []):
                            if str(g.get('sku_id', '')) == str(sku_id):
                                return g.get('quantity', 0)
            return 0
        except Exception:
            return None

    def _get_qfy_cart_qty_by_mid(self, supplier_name, mid):
        """直接使用 mid 查询庆丰裕购物车数量"""
        config = self.QFY_SYSTEMS.get(supplier_name)
        if not config or not mid:
            return None
        token = self._get_qfy_token(supplier_name)
        if not token:
            return None
        try:
            cached = self._qfy_tokens.get(supplier_name, {})
            sess = cached.get('session') or self.session
            cart_data_req = {'token': token, 'time': str(int(time.time() * 1000))}
            cart_resp = sess.post(
                f"{config['api_url']}/web_mycart.html",
                data=cart_data_req,
                timeout=15
            )
            cart_result = cart_resp.json()
            if cart_result.get('status') == 1:
                for group in cart_result.get('normal', []):
                    for item in group.get('list', []):
                        if str(item.get('mid', '')) == str(mid) or str(item.get('id', '')) == str(mid):
                            return item.get('buynum', 0)
                for item in cart_result.get('invalid', []):
                    if str(item.get('mid', '')) == str(mid) or str(item.get('id', '')) == str(mid):
                        return item.get('buynum', 0)
            return 0
        except Exception:
            return None

    # ===================== 通用方法 =====================

    def _error_result(self, supplier_name, barcode, error_msg):
        """生成错误结果"""
        url_map = {
            '俊龙': 'http://shop.szsjlyy.com/category?keyword=',
            '康之源': 'https://www.gdkzyyy.com/product/list.html?word=',
            '中源': 'http://shop.szszyyy.com/product/list.html?word=',
            '明华堂': 'https://web.szmhtyy.com/product/list.html?word=',
        }
        base = url_map.get(supplier_name, '')
        return {
            'status': 'error',
            'product_name': '没有',
            'specification': '-',
            'price': None,
            'original_price': None,
            'discount_price': None,
            'validity_date': '-',
            'stock_status': '-',
            'direct_link': base + barcode,
            'error': error_msg,
            'product_image': None,
            'manufacturer': None,
            'cart_quantity': None,
        }

    def _wrap_result(self, name, barcode, result):
        """包装搜索结果，添加供应商信息"""
        url_map = {
            '康之源': 'https://www.gdkzyyy.com',
            '明华堂': 'https://web.szmhtyy.com',
            '俊龙': 'http://shop.szsjlyy.com',
            '中源': 'http://shop.szszyyy.com',
        }
        result['supplier_name'] = name
        result['supplier_url'] = url_map.get(name, '')
        return result

    def search_all_suppliers(self, barcode):
        """搜索所有供应商 - 并行搜索 + 并行查询购物车数量"""
        results = [None] * 4
        supplier_names = ['康之源', '明华堂', '俊龙', '中源']
        search_methods = [
            lambda: self._wrap_result('康之源', barcode, self.search_kangzhiyuan(barcode)),
            lambda: self._wrap_result('明华堂', barcode, self.search_minghuatang(barcode)),
            lambda: self._wrap_result('俊龙', barcode, self.search_junlong(barcode)),
            lambda: self._wrap_result('中源', barcode, self.search_zhongyuan(barcode)),
        ]
        supplier_urls = {
            '康之源': 'https://www.gdkzyyy.com',
            '明华堂': 'https://web.szmhtyy.com',
            '俊龙': 'http://shop.szsjlyy.com',
            '中源': 'http://shop.szszyyy.com',
        }

        def do_search(idx):
            try:
                results[idx] = search_methods[idx]()
            except Exception as e:
                results[idx] = self._error_result(supplier_names[idx], barcode, str(e))

        threads = [threading.Thread(target=do_search, args=(i,)) for i in range(4)]
        for t in threads:
            t.start()
        for t in threads:
            t.join(timeout=20)

        # 过滤掉失败的结果（None表示超时）
        results = [r for r in results if r is not None]

        # 并行查询购物车数量
        def fetch_cart_qty(r):
            try:
                if r.get('status') != 'found':
                    return
                supplier = r['supplier_name']
                if supplier == '俊龙':
                    sku_id = r.pop('_sku_id', None)
                    if sku_id:
                        r['cart_quantity'] = self._get_junlong_cart_qty_by_skuid(sku_id)
                elif supplier in self.QFY_SYSTEMS:
                    mid = r.pop('_mid', None)
                    if mid:
                        r['cart_quantity'] = self._get_qfy_cart_qty_by_mid(supplier, mid)
            except Exception:
                pass

        threads = [threading.Thread(target=fetch_cart_qty, args=(r,)) for r in results]
        for t in threads:
            t.start()
        for t in threads:
            t.join(timeout=10)

        for r in results:
            r.pop('_sku_id', None)
            r.pop('_mid', None)

        return results


# ============================================================
# 模块级单例爬虫 - token跨请求缓存，避免每次搜索都重新登录
# ============================================================
_crawler_instance = None

def get_crawler():
    """获取全局爬虫单例（token缓存复用，大幅提升速度）"""
    global _crawler_instance
    if _crawler_instance is None:
        _crawler_instance = PharmacyCrawler()
    return _crawler_instance

# ============================================================
# API路由
# ============================================================

@app.route('/')
def index():
    """首页 - 返回前端页面"""
    return send_from_directory('static', 'index.html')

@app.route('/health')
def health():
    """健康检查"""
    return jsonify({'status': 'ok', 'timestamp': datetime.datetime.now().isoformat()})

# 供应商管理API
@app.route('/api/suppliers', methods=['GET'])
def get_suppliers():
    """获取所有供应商"""
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT * FROM suppliers WHERE status = 1")
    suppliers = cursor.fetchall()
    db.close()
    
    return jsonify({
        'success': True,
        'data': [dict(row) for row in suppliers]
    })

@app.route('/api/suppliers', methods=['POST'])
def add_supplier():
    """添加供应商"""
    data = request.get_json()
    
    db = get_db()
    cursor = db.cursor()
    cursor.execute('''
        INSERT INTO suppliers (name, url, account, password)
        VALUES (?, ?, ?, ?)
    ''', (data['name'], data['url'], data['account'], data['password']))
    db.commit()
    db.close()
    
    return jsonify({'success': True, 'message': '供应商添加成功'})

@app.route('/api/suppliers/<int:supplier_id>', methods=['PUT'])
def update_supplier(supplier_id):
    """更新供应商"""
    data = request.get_json()
    
    db = get_db()
    cursor = db.cursor()
    cursor.execute('''
        UPDATE suppliers 
        SET name = ?, url = ?, account = ?, password = ?
        WHERE id = ?
    ''', (data['name'], data['url'], data['account'], data['password'], supplier_id))
    db.commit()
    db.close()
    
    return jsonify({'success': True, 'message': '供应商更新成功'})

@app.route('/api/suppliers/<int:supplier_id>', methods=['DELETE'])
def delete_supplier(supplier_id):
    """删除供应商"""
    db = get_db()
    cursor = db.cursor()
    cursor.execute("UPDATE suppliers SET status = 0 WHERE id = ?", (supplier_id,))
    db.commit()
    db.close()
    
    return jsonify({'success': True, 'message': '供应商删除成功'})

# 药品搜索API
@app.route('/api/search', methods=['POST'])
def search_drugs():
    """搜索药品 - 只用条码"""
    data = request.get_json()
    barcode = data.get('barcode', '')
    
    if not barcode:
        return jsonify({'success': False, 'message': '请输入药品条码'})
    
    # 保存搜索记录
    db = get_db()
    cursor = db.cursor()
    cursor.execute('''
        INSERT INTO search_records (barcode, product_name, specification)
        VALUES (?, ?, ?)
    ''', (barcode, '', ''))
    search_record_id = cursor.lastrowid
    db.commit()
    
    # 执行爬虫搜索（包含并行购物车数量查询）
    crawler = get_crawler()
    results = crawler.search_all_suppliers(barcode)
    
    # 保存搜索结果
    for result in results:
        cursor.execute('''
            INSERT INTO price_results 
            (search_record_id, supplier_name, product_name, specification, price, validity_date, stock_status, direct_link, search_status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            search_record_id,
            result['supplier_name'],
            result['product_name'],
            result['specification'],
            result['price'],
            result['validity_date'],
            result['stock_status'],
            result['direct_link'],
            result['status']
        ))
    
    db.commit()
    db.close()
    
    return jsonify({
        'success': True,
        'data': {
            'search_id': search_record_id,
            'results': results
        }
    })

# 批量搜索API - 只解析Excel返回条码列表（前端逐个搜索）
@app.route('/api/search/batch', methods=['POST'])
def batch_search():
    """批量搜索 - 只解析Excel返回条码列表，前端逐个调用/api/search"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '请上传Excel文件'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '请选择文件'})

    try:
        # 读取Excel文件 - 统一使用openpyxl引擎，避免xlrd版本问题
        filename = file.filename.lower()
        try:
            if filename.endswith('.xls'):
                try:
                    df = pd.read_excel(file, engine='xlrd')
                except Exception:
                    file.seek(0)
                    df = pd.read_excel(file, engine='openpyxl')
            elif filename.endswith('.xlsx'):
                df = pd.read_excel(file, engine='openpyxl')
            else:
                df = pd.read_excel(file, engine='openpyxl')
        except Exception as read_err:
            return jsonify({'success': False, 'message': f'文件读取失败: {str(read_err)}。请确保上传的是有效的Excel文件(.xls或.xlsx)'})

        # 获取所有列名
        columns = list(df.columns)

        # 查找条码列
        barcode_col = None
        for col in columns:
            col_str = str(col).strip().replace('*', '')
            if col_str in ['条码', '条形码', '商品条码', 'barcode', '商品编码', '编码']:
                barcode_col = col
                break

        if not barcode_col:
            return jsonify({
                'success': False,
                'message': f'Excel文件格式错误，未找到条码列。当前列名：{columns}'
            })

        # 收集所有条码
        barcodes = []
        for index, row in df.iterrows():
            barcode = str(row[barcode_col]) if pd.notna(row[barcode_col]) else ''
            if barcode and barcode.lower() != 'nan':
                barcodes.append(barcode)

        if not barcodes:
            return jsonify({'success': False, 'message': '未找到有效的条码数据'})

        # 只返回条码列表，前端逐个搜索
        return jsonify({
            'success': True,
            'data': [{'barcode': bc} for bc in barcodes]
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'文件处理错误: {str(e)}'})

# 手动输入批量搜索API
@app.route('/api/search/manual-batch', methods=['POST'])
def manual_batch_search():
    """手动输入批量搜索 - 接收条码列表，返回条码数据供前端逐个搜索"""
    data = request.get_json()
    barcodes = data.get('barcodes', [])

    # 验证不超过20个
    if len(barcodes) > 20:
        return jsonify({'success': False, 'message': '手动输入最多支持20个条码'})

    # 验证每个条码不为空
    barcodes = [b.strip() for b in barcodes if b.strip()]
    if not barcodes:
        return jsonify({'success': False, 'message': '请输入至少一个条码'})

    return jsonify({'success': True, 'data': [{'barcode': bc} for bc in barcodes]})

# 仪表盘统计API
@app.route('/api/stats', methods=['GET'])
def get_stats():
    """获取仪表盘统计数据"""
    db = get_db()
    cursor = db.cursor()
    try:
        # 今日搜索次数
        cursor.execute("SELECT COUNT(*) FROM search_records WHERE DATE(created_at) = DATE('now', 'localtime')")
        today_count = cursor.fetchone()[0]

        # 已查询药品总数（去重条码）
        cursor.execute("SELECT COUNT(DISTINCT barcode) FROM search_records")
        total_drugs = cursor.fetchone()[0]

        # 搜索结果总数
        cursor.execute("SELECT COUNT(*) FROM price_results")
        total_results = cursor.fetchone()[0]

        cursor.close()
        return jsonify({
            'success': True,
            'data': {
                'today_search_count': today_count,
                'total_drugs': total_drugs,
                'saved_amount': 0,
                'total_results': total_results
            }
        })
    except Exception as e:
        cursor.close()
        return jsonify({'success': True, 'data': {'today_search_count': 0, 'total_drugs': 0, 'saved_amount': 0, 'total_results': 0}})

# 导出结果API - 支持分3条码一组导出
@app.route('/api/export', methods=['POST'])
def export_results():
    """导出比较结果为Excel，每3条码一组"""
    data = request.get_json()
    results = data.get('results', [])
    group_index = data.get('group_index', None)

    try:
        # 保存到当前工作目录下的exports文件夹 - 确保跨平台兼容
        export_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exports')
        os.makedirs(export_dir, exist_ok=True)

        # 收集导出数据
        export_data = []
        target_results = []
        if group_index is not None:
            start = group_index * 3
            end = min(start + 3, len(results))
            target_results = results[start:end]
            filename = f'pharmacy_export_group{group_index + 1}_{int(time.time())}.xlsx'
        else:
            target_results = results
            filename = f'pharmacy_export_{int(time.time())}.xlsx'

        for item in target_results:
            for supplier_result in item.get('supplier_results', []):
                price_val = supplier_result.get('price')
                export_data.append({
                    '条码': item.get('barcode', ''),
                    '供应商': supplier_result.get('supplier_name', ''),
                    '商品名称': supplier_result.get('product_name', '没有'),
                    '规格': supplier_result.get('specification', '没有'),
                    '原价(元)': supplier_result.get('original_price') or '-',
                    '折后价(元)': supplier_result.get('discount_price') or '-',
                    '价格(元)': price_val if price_val is not None else '-',
                    '有效期': supplier_result.get('validity_date', '没有'),
                    '库存状态': supplier_result.get('stock_status', '没有'),
                    '生产厂家': supplier_result.get('manufacturer', ''),
                    '直达链接': supplier_result.get('direct_link', '')
                })

        # 检查导出数据是否为空
        if not export_data:
            return jsonify({'success': False, 'message': '没有可导出的结果'}), 400

        # 使用openpyxl直接写入，避免pandas版本兼容问题
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "比价结果"

        # 写入表头
        if export_data:
            headers = list(export_data[0].keys())
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 写入数据
            for row_idx, row_data in enumerate(export_data, 2):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=row_data[header])
                    cell.alignment = Alignment(horizontal='left', vertical='center')

            # 自动调整列宽
            for col_idx, header in enumerate(headers, 1):
                max_length = len(str(header))
                for row_data in export_data:
                    max_length = max(max_length, len(str(row_data.get(header, ''))))
                ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A' + chr(64 + col_idx - 26)].width = min(max_length + 4, 60)

        filepath = os.path.join(export_dir, filename)
        wb.save(filepath)

        return send_from_directory(os.path.dirname(filepath), os.path.basename(filepath), as_attachment=True)

    except Exception as e:
        import traceback
        return jsonify({'success': False, 'message': f'导出失败: {str(e)}\n{traceback.format_exc()}'})

# ============================================================
# 购物车功能API - 直接加入供应商购物车
# 庆丰裕系统: web_addcart.html (mid + buynum)
# 俊龙: /api/order/cart (company_id + sku_id + quantity)
# ============================================================

@app.route('/api/cart/add', methods=['POST'])
def add_to_cart():
    """将商品直接加入供应商购物车
    请求参数:
    - supplier: 供应商名称 (俊龙/康之源/中源/明华堂)
    - barcode: 药品条码
    - quantity: 购买数量
    """
    data = request.get_json()
    supplier = data.get('supplier', '')
    barcode = data.get('barcode', '')
    quantity = int(data.get('quantity', 1))

    if not supplier or not barcode:
        return jsonify({'success': False, 'message': '缺少供应商或条码参数'})

    try:
        crawler = get_crawler()

        if supplier == '俊龙':
            # 俊龙加购: POST /api/order/cart
            token = crawler._get_junlong_token()
            if not token:
                return jsonify({'success': False, 'message': '俊龙登录失败，无法加入购物车'})

            session = requests.Session()
            session.headers.update({
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                'Accept': 'application/json',
                'Authorization': f'Bearer {token}',
                'Content-Type': 'application/json;charset=UTF-8',
                'Referer': 'http://shop.szsjlyy.com/',
            })

            # 搜索商品获取sku_id
            search_resp = session.get(
                'https://swoole.86yqy.com/api/shop/product/list',
                params={'keyword': barcode, 'page': 1, 'pageSize': 5},
                timeout=15
            )
            search_result = search_resp.json()
            items = search_result.get('data', [])

            if not items:
                return jsonify({'success': False, 'message': f'在{supplier}未找到该商品'})

            item = items[0]
            # 获取sku_id - 优先从sku数组取，否则用商品id
            sku_list = item.get('sku', [])
            if sku_list and len(sku_list) > 0:
                sku_id = sku_list[0].get('id', item.get('id'))
            else:
                sku_id = item.get('id')
            product_name = item.get('product_name', '')
            company_id = crawler.JUNLONG_COMPANY_ID

            # 加入购物车
            cart_data = {
                'company_id': company_id,
                'sku_id': sku_id,
                'quantity': quantity,
                'type': 0,
            }
            cart_resp = session.post(
                'https://swoole.86yqy.com/api/order/cart',
                json=cart_data,
                timeout=15
            )

            if cart_resp.status_code in (200, 201):
                # 加购成功后，重新查询购物车数量
                cart_qty = None
                try:
                    cart_qty = crawler._get_junlong_cart_qty_by_skuid(sku_id)
                except Exception:
                    pass
                return jsonify({
                    'success': True,
                    'message': f'已将 {product_name} x{quantity} 加入{supplier}购物车',
                    'data': {'supplier': supplier, 'product': product_name, 'quantity': quantity, 'cart_quantity': cart_qty}
                })
            else:
                return jsonify({'success': False, 'message': f'加入购物车失败({cart_resp.status_code}): {cart_resp.text[:200]}'})

        else:
            # 庆丰裕系统 (康之源/中源/明华堂): web_addcart.html
            config = crawler.QFY_SYSTEMS.get(supplier)
            if not config:
                return jsonify({'success': False, 'message': f'不支持的供应商: {supplier}'})

            token = crawler._get_qfy_token(supplier)
            if not token:
                return jsonify({'success': False, 'message': f'{supplier}登录失败，无法加入购物车'})

            cached = crawler._qfy_tokens.get(supplier, {})
            sess = cached.get('session') or requests.Session()

            # 搜索商品获取mid (商品ID)
            goods_data = {
                'token': token,
                'time': str(int(time.time() * 1000)),
                'param[word]': barcode,
                'order[val]': '1',
                'order[type]': 'desc',
                'hasqty': '0',
                'hasmedicare': '0',
                'buyhistory': '0',
                'hasstore': '0',
                'page': '1'
            }
            search_resp = sess.post(
                f"{config['api_url']}/web_goods_list.html",
                data=goods_data,
                timeout=15
            )
            search_result = search_resp.json()

            if search_result.get('status') != 1:
                return jsonify({'success': False, 'message': f'在{supplier}搜索商品失败'})

            items = search_result.get('data', {}).get('data', [])
            if not items:
                return jsonify({'success': False, 'message': f'在{supplier}未找到该商品'})

            mid = items[0].get('id')  # 商品规格ID
            product_name = items[0].get('goods_name', '')

            # 加入购物车 - 庆丰裕系统
            cart_data = {
                'token': token,
                'time': str(int(time.time() * 1000)),
                'mid': mid,
                'buynum': quantity,
            }
            cart_resp = sess.post(
                f"{config['api_url']}/web_addcart.html",
                data=cart_data,
                timeout=15
            )
            cart_result = cart_resp.json()

            if cart_result.get('status') == 1:
                # 加购成功后，重新查询购物车数量
                cart_qty = None
                try:
                    cart_qty = crawler._get_qfy_cart_qty_by_mid(supplier, mid)
                except Exception:
                    pass
                return jsonify({
                    'success': True,
                    'message': f'已将 {product_name} x{quantity} 加入{supplier}购物车',
                    'data': {'supplier': supplier, 'product': product_name, 'quantity': quantity, 'cart_quantity': cart_qty}
                })
            else:
                return jsonify({
                    'success': False,
                    'message': f'加入购物车失败: {cart_result.get("msg", "未知错误")}'
                })

    except Exception as e:
        import traceback
        return jsonify({'success': False, 'message': f'加入购物车出错: {str(e)}'})


# 购物车数量查询API - 查询某商品在购物车中的数量
@app.route('/api/cart/quantity', methods=['POST'])
def get_cart_quantity():
    """查询某条码商品在供应商购物车中的数量
    请求参数:
    - supplier: 供应商名称 (俊龙/康之源/中源/明华堂)
    - barcode: 药品条码
    """
    data = request.get_json()
    supplier = data.get('supplier', '')
    barcode = data.get('barcode', '')

    if not supplier or not barcode:
        return jsonify({'success': False, 'message': '缺少参数'})

    try:
        crawler = get_crawler()

        if supplier == '俊龙':
            token = crawler._get_junlong_token()
            if not token:
                return jsonify({'success': False, 'message': '登录失败'})

            session = requests.Session()
            session.headers.update({
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                'Accept': 'application/json',
                'Authorization': f'Bearer {token}',
                'Referer': 'http://shop.szsjlyy.com/',
            })

            # 先搜索获取sku_id
            search_resp = session.get(
                'https://swoole.86yqy.com/api/shop/product/list',
                params={'keyword': barcode, 'page': 1, 'pageSize': 5},
                timeout=15
            )
            search_result = search_resp.json()
            items = search_result.get('data', [])
            if not items:
                return jsonify({'success': True, 'quantity': 0})

            item = items[0]
            sku_list = item.get('sku', [])
            sku_id = sku_list[0].get('id', item.get('id')) if sku_list else item.get('id')

            # 获取购物车列表
            cart_resp = session.get(
                'https://swoole.86yqy.com/api/order/cart',
                params={'company_id': crawler.JUNLONG_COMPANY_ID},
                timeout=15
            )
            if cart_resp.status_code == 200:
                try:
                    cart_data = cart_resp.json()
                    # 数据结构: data[].data.valid_groups[].goods[]
                    for group in cart_data.get('data', []):
                        inner_data = group.get('data', {})
                        valid_groups = inner_data.get('valid_groups', [])
                        for vg in valid_groups:
                            for g in vg.get('goods', []):
                                if str(g.get('sku_id', '')) == str(sku_id):
                                    return jsonify({'success': True, 'quantity': g.get('quantity', 0)})
                except Exception as e:
                    logger.warning(f"俊龙购物车解析失败: {e}")
            return jsonify({'success': True, 'quantity': 0})

        else:
            # 庆丰裕系统 (康之源/中源/明华堂)
            config = crawler.QFY_SYSTEMS.get(supplier)
            if not config:
                return jsonify({'success': False, 'message': '不支持的供应商'})

            token = crawler._get_qfy_token(supplier)
            if not token:
                return jsonify({'success': False, 'message': '登录失败'})

            cached = crawler._qfy_tokens.get(supplier, {})
            sess = cached.get('session') or requests.Session()

            # 先搜索商品获取mid
            goods_data = {
                'token': token,
                'time': str(int(time.time() * 1000)),
                'param[word]': barcode,
                'order[val]': '1',
                'order[type]': 'desc',
                'hasqty': '0',
                'hasmedicare': '0',
                'buyhistory': '0',
                'hasstore': '0',
                'page': '1'
            }
            search_resp = sess.post(
                f"{config['api_url']}/web_goods_list.html",
                data=goods_data,
                timeout=15
            )
            search_result = search_resp.json()
            if search_result.get('status') != 1:
                return jsonify({'success': True, 'quantity': 0})

            items = search_result.get('data', {}).get('data', [])
            if not items:
                return jsonify({'success': True, 'quantity': 0})

            mid = items[0].get('id')

            # 修复: 使用 web_mycart.html 端点（与 /api/cart/list 保持一致）
            cart_data_req = {
                'token': token,
                'time': str(int(time.time() * 1000)),
            }
            cart_resp = sess.post(
                f"{config['api_url']}/web_mycart.html",
                data=cart_data_req,
                timeout=15
            )
            try:
                cart_result = cart_resp.json()
                if cart_result.get('status') == 1:
                    # 遍历 normal[].list[] 查找匹配的商品
                    for group in cart_result.get('normal', []):
                        for item in group.get('list', []):
                            if str(item.get('mid', '')) == str(mid) or str(item.get('id', '')) == str(mid):
                                return jsonify({'success': True, 'quantity': item.get('buynum', 0)})
                    # 也检查 invalid 列表
                    for item in cart_result.get('invalid', []):
                        if str(item.get('mid', '')) == str(mid) or str(item.get('id', '')) == str(mid):
                            return jsonify({'success': True, 'quantity': item.get('buynum', 0)})
            except Exception as e:
                logger.warning(f"{supplier}购物车列表解析失败: {e}")
            return jsonify({'success': True, 'quantity': 0})

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


# 购物车列表API - 获取供应商购物车全部内容
@app.route('/api/cart/list', methods=['POST'])
def get_cart_list():
    """获取供应商购物车列表"""
    data = request.get_json()
    supplier = data.get('supplier', '')

    if not supplier:
        return jsonify({'success': False, 'message': '缺少供应商参数'})

    try:
        crawler = get_crawler()

        if supplier == '俊龙':
            token = crawler._get_junlong_token()
            if not token:
                return jsonify({'success': False, 'message': '登录失败'})

            session = requests.Session()
            session.headers.update({
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                'Accept': 'application/json',
                'Authorization': f'Bearer {token}',
                'Referer': 'http://shop.szsjlyy.com/',
            })

            cart_resp = session.get(
                'https://swoole.86yqy.com/api/order/cart',
                params={'company_id': crawler.JUNLONG_COMPANY_ID},
                timeout=15
            )
            if cart_resp.status_code == 200:
                try:
                    cart_data = cart_resp.json()
                    # 数据结构: data[0].data.valid_groups[].goods[]
                    cart_items = []
                    for group in cart_data.get('data', []):
                        inner_data = group.get('data', {})
                        valid_groups = inner_data.get('valid_groups', [])
                        for vg in valid_groups:
                            for g in vg.get('goods', []):
                                sku = g.get('sku', {})
                                product = sku.get('product', {})
                                product_name = product.get('product_name', '') or product.get('standard_name', '')
                                if not product_name:
                                    continue  # 跳过无名商品
                                attr_value = sku.get('attr_value', '')
                                image = sku.get('image', '')
                                if image and not image.startswith('http'):
                                    image = 'https:' + image
                                cart_items.append({
                                    'id': g.get('uid', ''),
                                    'sku_id': g.get('sku_id', ''),
                                    'product_name': product_name,
                                    'specification': attr_value,
                                    'price': sku.get('whole_price', 0),
                                    'quantity': g.get('quantity', 0),
                                    'image': image,
                                    'stock': g.get('stock', 0),
                                    'stock_show': g.get('stock_show', ''),
                                    'expiry_date': g.get('expiry_date', '')
                                })
                    return jsonify({'success': True, 'data': cart_items})
                except Exception as e:
                    logger.warning(f"俊龙购物车解析失败: {e}")
            return jsonify({'success': True, 'data': []})

        else:
            config = crawler.QFY_SYSTEMS.get(supplier)
            if not config:
                return jsonify({'success': False, 'message': '不支持的供应商'})

            token = crawler._get_qfy_token(supplier)
            if not token:
                return jsonify({'success': False, 'message': '登录失败'})

            cached = crawler._qfy_tokens.get(supplier, {})
            sess = cached.get('session') or requests.Session()

            # 庆丰裕购物车列表 - 使用 web_mycart.html 端点
            # 数据结构: response.normal[].list[] 和 response.invalid[]
            cart_items = []
            try:
                cart_data = {
                    'token': token,
                    'time': str(int(time.time() * 1000)),
                }
                cart_resp = sess.post(
                    f"{config['api_url']}/web_mycart.html",
                    data=cart_data,
                    timeout=15
                )
                try:
                    cart_result = cart_resp.json()
                    if cart_result.get('status') == 1:
                        # normal商品（有效）
                        for group in cart_result.get('normal', []):
                            for item in group.get('list', []):
                                img = item.get('minimg') or item.get('goods_thumb') or item.get('thumb') or item.get('image') or ''
                                if img and not img.startswith('http'):
                                    # 庆丰裕系统图片处理：不同供应商图片后缀不同
                                    if supplier == '康之源':
                                        img = config['api_url'] + img + '_thumb.jpg'
                                    elif supplier == '明华堂':
                                        img = config['api_url'] + img + '.jpg'
                                    elif supplier == '中源':
                                        img = config['api_url'] + img + '.jpg'
                                cart_items.append({
                                    'id': item.get('id'),
                                    'product_name': item.get('goods_name', ''),
                                    'specification': item.get('specs', ''),
                                    'price': item.get('price', 0),
                                    'quantity': item.get('buynum', 0),
                                    'image': img,
                                    'stock': item.get('qty', 0),
                                    'validity_date': item.get('out_date', ''),
                                    'maker': item.get('maker', ''),
                                })
                        # invalid商品（失效/缺货）
                        for item in cart_result.get('invalid', []):
                            img = item.get('minimg') or item.get('goods_thumb') or item.get('thumb') or item.get('image') or ''
                            if img and not img.startswith('http'):
                                # 庆丰裕系统图片处理：不同供应商图片后缀不同
                                if supplier == '康之源':
                                    img = config['api_url'] + img + '_thumb.jpg'
                                elif supplier == '明华堂':
                                    img = config['api_url'] + img + '.jpg'
                                elif supplier == '中源':
                                    img = config['api_url'] + img + '.jpg'
                            cart_items.append({
                                'id': item.get('id'),
                                'product_name': item.get('goods_name', ''),
                                'specification': item.get('specs', ''),
                                'price': item.get('price', 0),
                                'quantity': item.get('buynum', 0),
                                'image': img,
                                'stock': item.get('qty', 0),
                                'validity_date': item.get('out_date', ''),
                                'maker': item.get('maker', ''),
                                'invalid': True,
                            })
                except ValueError:
                    logger.warning(f"{supplier}购物车列表返回非JSON: {cart_resp.text[:100]}")
            except Exception as e:
                logger.warning(f"{supplier}购物车列表获取失败: {e}")
            return jsonify({'success': True, 'data': cart_items})

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


# 购物车导出API - 将购物车商品导出为Excel
@app.route('/api/cart/export', methods=['POST'])
def export_cart():
    """导出购物车商品为Excel
    请求参数:
    - supplier: 供应商名称
    - items: 购物车商品列表 (可选，如果不传则自动从供应商获取)
    """
    data = request.get_json()
    supplier = data.get('supplier', '')
    items = data.get('items', [])

    try:
        export_data = []

        # 如果前端传了items，直接使用；否则自动获取购物车列表
        if items:
            for item in items:
                export_data.append({
                    '商品名称': item.get('product_name', ''),
                    '规格': item.get('specification', ''),
                    '价格(元)': item.get('price', 0),
                    '数量': item.get('quantity', 0),
                    '库存': item.get('stock', 0),
                    '有效期': item.get('validity_date', '') or item.get('expiry_date', ''),
                    '生产厂家': item.get('maker', '') or item.get('manufacturer', ''),
                    '商品图片': item.get('image', ''),
                })
        else:
            # 修复: 自动从供应商获取购物车 - 使用与 /api/cart/list 完全一致的逻辑
            if not supplier:
                return jsonify({'success': False, 'message': '缺少供应商参数'})
            crawler = get_crawler()

            if supplier == '俊龙':
                token = crawler._get_junlong_token()
                if not token:
                    return jsonify({'success': False, 'message': '登录失败'})
                session = requests.Session()
                session.headers.update({
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                    'Accept': 'application/json',
                    'Authorization': f'Bearer {token}',
                    'Referer': 'http://shop.szsjlyy.com/',
                })
                cart_resp = session.get(
                    'https://swoole.86yqy.com/api/order/cart',
                    params={'company_id': crawler.JUNLONG_COMPANY_ID},
                    timeout=15
                )
                if cart_resp.status_code == 200:
                    try:
                        cart_json = cart_resp.json()
                        for group in cart_json.get('data', []):
                            inner_data = group.get('data', {})
                            valid_groups = inner_data.get('valid_groups', [])
                            for vg in valid_groups:
                                for g in vg.get('goods', []):
                                    sku = g.get('sku', {})
                                    product = sku.get('product', {})
                                    product_name = product.get('product_name', '') or product.get('standard_name', '')
                                    if not product_name:
                                        continue
                                    image = sku.get('image', '')
                                    if image and not image.startswith('http'):
                                        image = 'https:' + image
                                    export_data.append({
                                        '商品名称': product_name,
                                        '规格': sku.get('attr_value', ''),
                                        '价格(元)': sku.get('whole_price', 0),
                                        '数量': g.get('quantity', 0),
                                        '库存': g.get('stock', 0),
                                        '有效期': g.get('expiry_date', ''),
                                        '生产厂家': '',
                                        '商品图片': image,
                                    })
                    except Exception as e:
                        logger.warning(f"俊龙购物车解析失败: {e}")
            else:
                # 庆丰裕系统 - 修复: 使用 web_mycart.html 端点（与 /api/cart/list 一致）
                config = crawler.QFY_SYSTEMS.get(supplier)
                if not config:
                    return jsonify({'success': False, 'message': '不支持的供应商'})
                token = crawler._get_qfy_token(supplier)
                if not token:
                    return jsonify({'success': False, 'message': '登录失败'})
                cached = crawler._qfy_tokens.get(supplier, {})
                sess = cached.get('session') or requests.Session()
                cart_data_req = {
                    'token': token,
                    'time': str(int(time.time() * 1000)),
                }
                cart_resp = sess.post(
                    f"{config['api_url']}/web_mycart.html",
                    data=cart_data_req,
                    timeout=15
                )
                try:
                    cart_result = cart_resp.json()
                    if cart_result.get('status') == 1:
                        for group in cart_result.get('normal', []):
                            for item in group.get('list', []):
                                img = item.get('minimg') or item.get('goods_thumb') or item.get('thumb') or item.get('image') or ''
                                if img and not img.startswith('http'):
                                    if supplier == '康之源':
                                        img = config['api_url'] + img + '_thumb.jpg'
                                    else:
                                        img = config['api_url'] + img + '.jpg'
                                export_data.append({
                                    '商品名称': item.get('goods_name', ''),
                                    '规格': item.get('specs', ''),
                                    '价格(元)': item.get('price', 0),
                                    '数量': item.get('buynum', 0),
                                    '库存': item.get('qty', 0),
                                    '有效期': item.get('out_date', ''),
                                    '生产厂家': item.get('maker', ''),
                                    '商品图片': img,
                                })
                        for item in cart_result.get('invalid', []):
                            img = item.get('minimg') or item.get('goods_thumb') or item.get('thumb') or item.get('image') or ''
                            if img and not img.startswith('http'):
                                if supplier == '康之源':
                                    img = config['api_url'] + img + '_thumb.jpg'
                                else:
                                    img = config['api_url'] + img + '.jpg'
                            export_data.append({
                                '商品名称': item.get('goods_name', ''),
                                '规格': item.get('specs', ''),
                                '价格(元)': item.get('price', 0),
                                '数量': item.get('buynum', 0),
                                '库存': item.get('qty', 0),
                                '有效期': item.get('out_date', ''),
                                '生产厂家': item.get('maker', ''),
                                '商品图片': img,
                            })
                except Exception as e:
                    logger.warning(f"{supplier}购物车解析失败: {e}")

        if not export_data:
            return jsonify({'success': False, 'message': '购物车为空，无数据可导出'})

        # 使用openpyxl直接写入
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "购物车"

        headers = list(export_data[0].keys())
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_idx, row_data in enumerate(export_data, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data[header])
                cell.alignment = Alignment(horizontal='left', vertical='center')

        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row_data in export_data:
                max_length = max(max_length, len(str(row_data.get(header, ''))))
            col_letter = chr(64 + col_idx) if col_idx <= 26 else 'A' + chr(64 + col_idx - 26)
            ws.column_dimensions[col_letter].width = min(max_length + 4, 60)

        export_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exports')
        os.makedirs(export_dir, exist_ok=True)
        filepath = os.path.join(export_dir, f'cart_export_{supplier}_{int(time.time())}.xlsx')
        wb.save(filepath)

        return send_from_directory(os.path.dirname(filepath), os.path.basename(filepath), as_attachment=True)

    except Exception as e:
        import traceback
        return jsonify({'success': False, 'message': f'导出失败: {str(e)}\n{traceback.format_exc()}'})


# 搜索历史API
@app.route('/api/search/history', methods=['GET'])
def search_history():
    """获取搜索历史"""
    db = get_db()
    cursor = db.cursor()
    cursor.execute('''
        SELECT sr.*, COUNT(pr.id) as result_count
        FROM search_records sr
        LEFT JOIN price_results pr ON sr.id = pr.search_record_id
        GROUP BY sr.id
        ORDER BY sr.created_at DESC
        LIMIT 50
    ''')
    records = cursor.fetchall()
    db.close()
    
    return jsonify({
        'success': True,
        'data': [dict(row) for row in records]
    })

# 搜索历史导出API
@app.route('/api/search/history/export', methods=['GET'])
def export_search_history():
    """导出搜索历史为Excel，包含每条记录的详细比价结果"""
    try:
        db = get_db()
        cursor = db.cursor()

        # 获取所有搜索记录
        cursor.execute('''
            SELECT sr.id, sr.barcode, sr.created_at
            FROM search_records sr
            ORDER BY sr.created_at DESC
        ''')
        records = cursor.fetchall()

        # 获取所有比价结果，按search_record_id分组
        cursor.execute('''
            SELECT search_record_id, supplier_name, product_name, specification,
                   price, validity_date, stock_status, direct_link, search_status
            FROM price_results
            ORDER BY search_record_id, id
        ''')
        all_results = cursor.fetchall()
        db.close()

        # 按search_record_id分组比价结果
        results_map = {}
        for r in all_results:
            sid = r['search_record_id']
            if sid not in results_map:
                results_map[sid] = []
            results_map[sid].append(dict(r))

        # 使用openpyxl直接写入
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "搜索历史"

        headers = ['条码', '搜索时间', '供应商', '商品名称', '规格', '价格(元)', '有效期', '库存状态', '直达链接', '搜索状态']
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        row_idx = 2
        for record in records:
            barcode = record['barcode']
            created_at = record['created_at']
            sid = record['id']
            pr_list = results_map.get(sid, [])

            if not pr_list:
                # 没有比价结果，仍然导出搜索记录
                ws.cell(row=row_idx, column=1, value=barcode)
                ws.cell(row=row_idx, column=2, value=created_at)
                ws.cell(row=row_idx, column=3, value='')
                ws.cell(row=row_idx, column=4, value='')
                ws.cell(row=row_idx, column=5, value='')
                ws.cell(row=row_idx, column=6, value='')
                ws.cell(row=row_idx, column=7, value='')
                ws.cell(row=row_idx, column=8, value='')
                ws.cell(row=row_idx, column=9, value='')
                ws.cell(row=row_idx, column=10, value='')
                row_idx += 1
            else:
                for i, pr in enumerate(pr_list):
                    ws.cell(row=row_idx, column=1, value=barcode if i == 0 else '')
                    ws.cell(row=row_idx, column=2, value=created_at if i == 0 else '')
                    ws.cell(row=row_idx, column=3, value=pr.get('supplier_name', ''))
                    ws.cell(row=row_idx, column=4, value=pr.get('product_name', ''))
                    ws.cell(row=row_idx, column=5, value=pr.get('specification', ''))
                    price_val = pr.get('price')
                    ws.cell(row=row_idx, column=6, value=price_val if price_val is not None else '-')
                    ws.cell(row=row_idx, column=7, value=pr.get('validity_date', ''))
                    ws.cell(row=row_idx, column=8, value=pr.get('stock_status', ''))
                    ws.cell(row=row_idx, column=9, value=pr.get('direct_link', ''))
                    ws.cell(row=row_idx, column=10, value=pr.get('search_status', ''))
                    row_idx += 1

        # 调整列宽
        col_widths = {'A': 18, 'B': 25, 'C': 12, 'D': 30, 'E': 20, 'F': 12, 'G': 15, 'H': 12, 'I': 50, 'J': 12}
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        export_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exports')
        os.makedirs(export_dir, exist_ok=True)
        filepath = os.path.join(export_dir, f'search_history_{int(time.time())}.xlsx')
        wb.save(filepath)

        return send_from_directory(os.path.dirname(filepath), os.path.basename(filepath), as_attachment=True)

    except Exception as e:
        import traceback
        return jsonify({'success': False, 'message': f'导出失败: {str(e)}\n{traceback.format_exc()}'})

# ============================================================
# 图片代理API - 解决跨域Referer问题
# ============================================================

@app.route('/api/proxy/image')
def proxy_image():
    """代理加载供应商图片，解决跨域和Referer问题"""
    url = request.args.get('url', '')
    if not url:
        return jsonify({'success': False, 'message': '缺少url参数'}), 400

    try:
        # 确定Referer
        referer = None
        if 'gdkzyyy.com' in url:
            referer = 'https://www.gdkzyyy.com/'
        elif 'szszyyy.com' in url:
            referer = 'http://shop.szszyyy.com/'
        elif 'szmhtyy.com' in url:
            referer = 'https://web.szmhtyy.com/'
        elif 'szsjlyy.com' in url or '86yqy.com' in url:
            referer = 'http://shop.szsjlyy.com/'

        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'image/webp,image/apng,image/*,*/*;q=0.8',
        }
        if referer:
            headers['Referer'] = referer

        resp = requests.get(url, headers=headers, timeout=10, stream=True)
        if resp.status_code == 200:
            content_type = resp.headers.get('Content-Type', 'image/jpeg')
            return Response(resp.iter_content(chunk_size=8192),
                          content_type=content_type,
                          headers={'Cache-Control': 'max-age=3600'})
        else:
            return jsonify({'success': False, 'message': f'图片获取失败: {resp.status_code}'}), resp.status_code
    except Exception as e:
        logger.error(f"图片代理失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================
# 主入口
# ============================================================

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='127.0.0.1', port=port, debug=False)
